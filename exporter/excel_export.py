"""
exporter/excel_export.py

Generates the Excel summary workbook from a list of parsed UPS invoice dicts
(output of parser.ups_parser.parse_invoice).

One sheet (Sheet1) with two tables:
  Table 1 — invoice summary (12 columns, A–L)
  Table 2 — charge breakdown (21 columns, A–U), placed below Table 1

Designed for the new snake_case field names from ups_parser.py.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Style helpers ─────────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Arial", bold=bold, size=size, color="000000")


def _no_fill() -> PatternFill:
    return PatternFill(fill_type=None)


def _grey_fill() -> PatternFill:
    return PatternFill("solid", start_color="F2F2F2", fgColor="F2F2F2")


def _cell(ws, row: int, col: int, value: Any = "", bold: bool = False,
          size: int = 10, fill=None, ha: str = "left", wrap: bool = False,
          fmt: str | None = None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(bold, size)
    c.fill = fill if fill is not None else _no_fill()
    c.border = _thin_border()
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    return c


NUM = "#,##0.00"
TXT = "@"

# ── Sort key ──────────────────────────────────────────────────────────────────

def _sort_key(inv: dict):
    """Sort all invoices by invoice date ascending, regardless of account."""
    raw_date = inv.get("invoice_date", "")
    try:
        return datetime.strptime(raw_date, "%B %d, %Y")
    except ValueError:
        return datetime.min


def _safe(inv: dict, key: str, default: Any = 0.0) -> Any:
    v = inv.get(key, default)
    return v if v is not None else default


# ── Main workbook builder ─────────────────────────────────────────────────────

def build_workbook(invoices: list[dict]) -> openpyxl.Workbook:
    """
    Build and return the summary workbook.

    Accepts invoices in the snake_case format from parse_invoice().
    
    """
    if not invoices:
        raise ValueError("No invoices to export")

    sorted_invs = sorted(invoices, key=_sort_key)

    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True
    ws = wb.active
    ws.title = "Sheet1"

    # ── TABLE 1: Invoice Summary ──────────────────────────────────────────────

    T1_HEADERS = [
        "Invoice Number", "Account Number", "Amount Due", "Invoice Date",
        "Invoice Status", "Payment Status", "Subtotal", "Tax",
        "Government Charges", "Billed Amount", "Due Date", "Type",
    ]
    T1_WIDTHS   = [22, 16, 12, 20, 14, 16, 14, 10, 20, 14, 20, 18]
    T1_NUM_COLS = {7, 8, 9, 10}   # G H I J — currency

    for i, w in enumerate(T1_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    T1_HDR_ROW = 1
    ws.row_dimensions[T1_HDR_ROW].height = 30
    for i, h in enumerate(T1_HEADERS, 1):
        _cell(ws, T1_HDR_ROW, i, h, bold=True, ha="center", wrap=True)

    T1_DATA_START = 2
    for row_offset, inv in enumerate(sorted_invs):
        r = T1_DATA_START + row_offset
        ws.row_dimensions[r].height = 15

        row_vals = [
            str(_safe(inv, "invoice_number", "")),   # A — text, leading zeros
            _safe(inv, "account_number", ""),         # B
            _safe(inv, "amount_due", "$0.00"),        # C
            _safe(inv, "invoice_date", ""),           # D — text
            _safe(inv, "invoice_status", "Closed"),   # E
            _safe(inv, "payment_status", "Accepted"), # F
            None,                                     # G — formula written below
            _safe(inv, "tax", 0.0),                   # H
            _safe(inv, "government_charges", 0.0),    # I
            _safe(inv, "billed_amount", 0.0),         # J
            _safe(inv, "due_date", ""),               # K — text
            _safe(inv, "invoice_type_label", ""),     # L
        ]

        for col, val in enumerate(row_vals, 1):
            if col in T1_NUM_COLS and col != 7:
                _cell(ws, r, col, val, fmt=NUM, ha="right")
            elif col in (1, 4, 11):
                # BUG-03: pass fmt=TXT directly into _cell so the @ format is
                # applied in the same call that sets the value. Setting it
                # afterwards risks Excel reinterpreting the value first.
                _cell(ws, r, col, val, fmt=TXT, ha="left")
            else:
                _cell(ws, r, col, val, ha="left")

        # Column G — Subtotal: MUST be an Excel formula, never a computed float
        print(f"Writing subtotal formula to G{r}: =J{r}-I{r}-H{r}", flush=True)
        sc = ws.cell(row=r, column=7, value=f"=J{r}-I{r}-H{r}")
        sc.font = _font()
        sc.fill = _no_fill()
        sc.border = _thin_border()
        sc.number_format = NUM
        sc.alignment = Alignment(horizontal="right", vertical="center")
        print(f"G{r} cell value after write: {sc.value!r}", flush=True)

    T1_LAST_DATA  = T1_DATA_START + len(sorted_invs) - 1
    T1_TOTALS_ROW = T1_LAST_DATA + 1
    ws.row_dimensions[T1_TOTALS_ROW].height = 15
    for col in range(1, 13):
        if col in T1_NUM_COLS:
            cl = get_column_letter(col)
            _cell(ws, T1_TOTALS_ROW, col,
                  f"=SUM({cl}{T1_DATA_START}:{cl}{T1_LAST_DATA})",
                  bold=True, fill=_grey_fill(), ha="right", fmt=NUM)
        else:
            c = ws.cell(row=T1_TOTALS_ROW, column=col, value="")
            c.fill = _grey_fill()
            c.border = _thin_border()
            c.font = _font(bold=True)

    # Per-account breakdown rows
    ACCT_HDR_ROW  = T1_TOTALS_ROW + 2
    ACCT_4172_ROW = T1_TOTALS_ROW + 3
    ACCT_Y8_ROW   = T1_TOTALS_ROW + 4
    for rr in (ACCT_HDR_ROW, ACCT_4172_ROW, ACCT_Y8_ROW):
        ws.row_dimensions[rr].height = 15

    for col, lbl in [(7, "Subtotal"), (8, "Tax"), (9, "Govt Charges"), (10, "Billed Total")]:
        _cell(ws, ACCT_HDR_ROW, col, lbl, bold=True, ha="center")

    for r, acct in [(ACCT_4172_ROW, "4172AV"), (ACCT_Y8_ROW, "Y8A864")]:
        group = [i for i in sorted_invs
                 if str(i.get("account_number", "")).upper() == acct]
        _cell(ws, r, 6, acct, bold=True, ha="left")
        sub = sum(
            _safe(i, "billed_amount") - _safe(i, "government_charges") - _safe(i, "tax")
            for i in group
        )
        tax = sum(_safe(i, "tax") for i in group)
        gov = sum(_safe(i, "government_charges") for i in group)
        bil = sum(_safe(i, "billed_amount") for i in group)
        for col, val in [(7, sub), (8, tax), (9, gov), (10, bil)]:
            _cell(ws, r, col, round(val, 2), bold=True, ha="right", fmt=NUM)

    # ── TABLE 2: Charge Breakdown ─────────────────────────────────────────────

    T2_START_ROW = ACCT_Y8_ROW + 4
    T2_HEADERS = [
        "Invoice Number", "Account Number",
        "Import Freight", "Fuel Surcharge", "Print Label", "Surge Fees",
        "Discounts Applied", "Worldwide Service", "UPS CampusShip",
        "UPS Returns", "Adjustments & Other Charges", "Service Charges (Weekly)",
        "Govt Agency Fee", "Additional Tariff Line Fee", "Brokerage GST/HST",
        "Duty (US Customs)", "Merchandise Processing Fee", "Disbursement Fee",
        "Entry Prep Fee", "PGA Disclaim Fee", "Total Charges",
    ]
    T2_WIDTHS = [22, 16, 16, 16, 12, 12, 18, 18, 16, 14, 26, 22,
                 16, 24, 18, 18, 24, 18, 14, 16, 14]

    # Set column widths (take max of T1 and T2 for shared columns 1-12)
    for i, w in enumerate(T2_WIDTHS, 1):
        col_l = get_column_letter(i)
        current = ws.column_dimensions[col_l].width or 0
        if w > current:
            ws.column_dimensions[col_l].width = w
    for i in range(13, 22):
        ws.column_dimensions[get_column_letter(i)].width = T2_WIDTHS[i - 1]

    ws.row_dimensions[T2_START_ROW].height = 30
    for i, h in enumerate(T2_HEADERS, 1):
        _cell(ws, T2_START_ROW, i, h, bold=True, ha="center", wrap=True)

    T2_CHARGE_KEYS = [
        "import_freight", "fuel_surcharge", "print_label", "surge_fees",
        "discounts_applied", "worldwide_service", "campus_ship", "ups_returns",
        "adjustments_other", "service_charges", "govt_agency_fee", "tariff_line_fee",
        "brokerage_gst", "duty_us", "merch_proc_fee", "disbursement_fee",
        "entry_prep_fee", "pga_disclaim_fee",
    ]

    T2_DATA_START = T2_START_ROW + 1
    for row_offset, inv in enumerate(sorted_invs):
        r = T2_DATA_START + row_offset
        ws.row_dimensions[r].height = 15

        # BUG-03: fmt=TXT passed inline so @ format is set with the value in one step
        _cell(ws, r, 1, str(_safe(inv, "invoice_number", "")), fmt=TXT, ha="left")
        _cell(ws, r, 2, _safe(inv, "account_number", ""), ha="left")

        for col_offset, key in enumerate(T2_CHARGE_KEYS):
            col = col_offset + 3
            _cell(ws, r, col, _safe(inv, key, 0.0), ha="right", fmt=NUM)

        _cell(ws, r, 21, _safe(inv, "billed_amount", 0.0), ha="right", fmt=NUM)

    T2_LAST_DATA  = T2_DATA_START + len(sorted_invs) - 1
    T2_TOTALS_ROW = T2_LAST_DATA + 1
    ws.row_dimensions[T2_TOTALS_ROW].height = 15
    for col in range(1, 22):
        if col >= 3:
            cl = get_column_letter(col)
            _cell(ws, T2_TOTALS_ROW, col,
                  f"=SUM({cl}{T2_DATA_START}:{cl}{T2_LAST_DATA})",
                  bold=True, fill=_grey_fill(), ha="right", fmt=NUM)
        else:
            c = ws.cell(row=T2_TOTALS_ROW, column=col, value="")
            c.fill = _grey_fill()
            c.border = _thin_border()
            c.font = _font(bold=True)

    ws.freeze_panes = "A2"
    return wb


def build_workbook_bytes(invoices: list[dict]) -> bytes:
    """Build the two-table workbook and return raw .xlsx bytes."""
    wb = build_workbook(invoices)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# Alias kept for any code still referencing the old name
build_summary_workbook_bytes = build_workbook_bytes
